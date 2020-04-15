import os

from util import get_newest,readImage,one_hot,average,parse_metric_txt
from arg_parser import arg_parser
from openpyxl import load_workbook,Workbook


a = arg_parser()

model_dict = {'unet': 'U-Net', 'r2u': 'R2U-Net', 'att': 'Attention U-Net', 'uplus': 'UNet++', 'ce': 'CE-Net', 'vnet': 'V-Net', \
              'zigzag_re': 'ZigZag-U-Net-residual', 'zigzag_re_r': 'ZigZag-U-Net-residual-reverse', 'zigzag_r': 'ZigZag-U-Net-regular', 'zigzag_r_r': 'ZigZag-U-Net-regular-reverse', \
              'zigzag_d': 'ZigZag-U-Net-dense', 'attv': 'Attention V-Net'}
target_dict = {'HN_OAR': 'HaN_OAR', 'Lu_OAR': 'Thoracic_OAR', 'HN_GTV': 'Naso_GTV', 'Lu_GTV': 'Lung_GTV'}
num_class_dict = {'HaN_OAR': 23, 'Thoracic_OAR': 7, 'Lung_GTV': 2, 'Naso_GTV': 2}

a.add_map('--model', model_dict)
a.add_map('--target', target_dict)

ret_dict = a()

model_key = ret_dict['--model']
target = ret_dict['--target']

work_path = get_newest(os.path.join('build', '{}-{}'.format(model_key, target), 'test_result'))
pb_name = os.path.split(work_path)[1]
work_process_path = os.path.join('build', '{}-{}'.format(model_key, target), 'test_result_process', pb_name)

metric_list = ['DSC']

def do(metric, extra_name, model_key, target, base_path):
    analysis_path = 'build/{}{}-analysis.xlsx'.format(metric, extra_name)

    if(os.path.exists(analysis_path)):
        wb = load_workbook(analysis_path)
    else:
        wb = Workbook()

    ws = wb.create_sheet('{}-{}'.format(model_key, target))

    work_file_path = os.path.join(base_path, '{}.txt'.format(metric))
    evaluate_result_dict = parse_metric_txt(work_file_path)

    key_list = list(evaluate_result_dict.keys())
    [ws.cell(1, i + 1, key_list[i]) for i in range(len(key_list))]
    for i in range(len(key_list)):
        key = key_list[i]  
        target_result = evaluate_result_dict[key]  
        for j in range(len(target_result)):
            ws.cell(j + 2, i + 1).value = target_result[j]
    
    wb.save(analysis_path)

for metric in metric_list:
    do(metric, '', model_key, target, work_path)
    do(metric, '-process', model_key, target, work_process_path)